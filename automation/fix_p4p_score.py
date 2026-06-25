"""
fix_p4p_score.py
----------------
Generalized fixer for P4P physician score Excel files.

Strategy:
  1. Find the "score column" — the rightmost column whose header matches a
     score keyword (รวมแต้ม / รวมคะแนน / คะแนนรวม / คะแนน / แต้ม).
  2. Find the "total anchor cell" — any cell whose text contains a grand-total
     keyword (รวมทั้งหมด / รวมคะแนน / รวมแต้มทั้งหมด / คะแนนรวม) or that
     embeds a large number after such a keyword.
  3. Determine the data range of the score column (first numeric row → row
     just above the total anchor).
  4. Insert =SUM(...) in the score column at the total anchor's row, then
     update the anchor cell to reference the formula cell dynamically.
  5. Save in-place (or to an output path if provided).

Usage:
    python fix_p4p_score.py <excel_file> [output_file]

Example:
    python fix_p4p_score.py "P4P ธีรุตม์.xlsx"
    python fix_p4p_score.py "P4P สมชาย.xlsx" "P4P สมชาย_fixed.xlsx"
"""

import sys
import io
import re
import openpyxl

# Force UTF-8 output so Thai characters print correctly on Windows (cp874 terminal)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string


# ── keyword lists ────────────────────────────────────────────────────────────
# Keywords that signal a grand-total label
TOTAL_KEYWORDS = [
    'รวมแต้มทั้งหมด', 'รวมคะแนนทั้งหมด', 'คะแนนรวมทั้งหมด',
    'รวมทั้งหมด', 'รวมคะแนน', 'คะแนนรวม',
]

# Keywords used as column headers for the individual score column
SCORE_HEADER_KEYWORDS = ['รวมแต้ม', 'รวมคะแนน', 'คะแนนรวม', 'คะแนน', 'แต้ม']

# How far (rows/cols) to search around an anchor for a numeric value
SEARCH_RADIUS = 6


# ── helpers ──────────────────────────────────────────────────────────────────

def cell_text(cell):
    """Return stripped string value or '' if not a string."""
    return cell.value.strip() if isinstance(cell.value, str) else ''


def extract_number_from_text(text):
    """Pull the first large float / int out of a string (ignores small values)."""
    for m in re.finditer(r'[\d,]+\.?\d*', text.replace(',', '')):
        try:
            n = float(m.group())
            if n > 100:          # ignore day numbers like 1–31
                return n
        except ValueError:
            pass
    return None


def find_score_column(ws):
    """
    Return (col_index, header_row) for the rightmost column whose header
    matches a SCORE_HEADER_KEYWORD.
    Prefers more specific keywords (longer match first).
    """
    best = None           # (specificity, col, row, coord)
    for row in ws.iter_rows():
        for cell in row:
            txt = cell_text(cell)
            if not txt:
                continue
            for kw in SCORE_HEADER_KEYWORDS:
                if kw in txt:
                    spec = len(kw)
                    # prefer rightmost, then most specific
                    if best is None or cell.column > best[1] or (
                            cell.column == best[1] and spec > best[0]):
                        best = (spec, cell.column, cell.row, cell.coordinate)
    if best:
        return best[1], best[2]
    return None, None


def find_total_anchor(ws):
    """
    Return the cell that acts as the grand-total label.
    Searches for TOTAL_KEYWORDS; returns the cell with the most specific match.
    """
    best = None           # (specificity, row, col, cell)
    for row in ws.iter_rows():
        for cell in row:
            txt = cell_text(cell)
            if not txt:
                continue
            for kw in TOTAL_KEYWORDS:
                if kw in txt:
                    spec = len(kw)
                    if best is None or spec > best[0]:
                        best = (spec, cell.row, cell.column, cell)
    return best[3] if best else None


def nearest_number_near_cell(ws, anchor_row, anchor_col, score_col):
    """
    Search nearby cells for a numeric value (or a text cell embedding one).
    Prioritise cells in the score column first, then general proximity.
    """
    candidates = []
    for dr in range(-SEARCH_RADIUS, SEARCH_RADIUS + 1):
        for dc in range(-SEARCH_RADIUS, SEARCH_RADIUS + 1):
            r = anchor_row + dr
            c = anchor_col + dc
            if r < 1 or c < 1:
                continue
            cell = ws.cell(r, c)
            dist = abs(dr) + abs(dc)
            # bonus: same column as score col
            col_bonus = 0 if c == score_col else 2
            val = cell.value
            if isinstance(val, (int, float)):
                candidates.append((dist + col_bonus, val, cell.coordinate))
            elif isinstance(val, str):
                n = extract_number_from_text(val)
                if n is not None:
                    candidates.append((dist + col_bonus + 1, n, cell.coordinate))
    candidates.sort()
    return candidates[0] if candidates else None


def get_score_data_range(ws, score_col, header_row, anchor_row):
    """
    Return (first_data_row, last_data_row) — all rows in score_col between
    header_row+1 and anchor_row-1 that could hold numeric scores.
    We set the range to header_row+1 .. anchor_row-1 and let SUM skip text.
    """
    first = header_row + 1
    last = anchor_row - 1
    return first, last


def unmerge_if_needed(ws, coordinate):
    """If the cell is inside a merged range, unmerge that range and return it."""
    for merged in list(ws.merged_cells.ranges):
        cell_obj = ws[coordinate]
        if (merged.min_row <= cell_obj.row <= merged.max_row and
                merged.min_col <= cell_obj.column <= merged.max_col):
            merge_str = str(merged)
            ws.unmerge_cells(merge_str)
            return merge_str
    return None


# ── main fixer ───────────────────────────────────────────────────────────────

def fix_p4p_file(input_path, output_path=None):
    if output_path is None:
        output_path = input_path

    print(f"\n{'='*60}")
    print(f"Processing: {input_path}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(input_path)

    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")

        # ── Step 1: Find score column ────────────────────────────────────────
        score_col, header_row = find_score_column(ws)
        if score_col is None:
            print("  [SKIP] No score column header found.")
            continue
        col_letter = get_column_letter(score_col)
        print(f"  Score column : {col_letter} (col {score_col}), header at row {header_row}")

        # ── Step 2: Find total anchor ────────────────────────────────────────
        anchor = find_total_anchor(ws)
        if anchor is None:
            print("  [SKIP] No grand-total keyword cell found.")
            continue
        anchor_display = anchor.value.strip()[:60] if anchor.value else ''
        print(f"  Total anchor : {anchor.coordinate}  ->  \"{anchor_display}\"")

        # ── Step 3: Verify current total (read data_only for computed values) ──
        wb_do = openpyxl.load_workbook(input_path, data_only=True)
        ws_do = wb_do[sheet_name]
        current = nearest_number_near_cell(ws_do, anchor.row, anchor.column, score_col)
        # Direct sum of data rows only (header and anchor rows excluded to avoid
        # double-counting sub-totals or the grand-total formula itself)
        direct_sum = sum(
            ws_do.cell(r, score_col).value
            for r in range(header_row + 1, anchor.row)
            if isinstance(ws_do.cell(r, score_col).value, (int, float))
        )
        print(f"  Direct SUM of score column = {direct_sum}")
        if current:
            print(f"  Existing total near anchor : {current[1]}  (at {current[2]})")

        # ── Step 4: Determine data range ─────────────────────────────────────
        first_row, last_row = get_score_data_range(
            ws, score_col, header_row, anchor.row)
        sum_range = f"{col_letter}{first_row}:{col_letter}{last_row}"
        sum_formula = f"=SUM({sum_range})"
        sum_cell_addr = f"{col_letter}{anchor.row}"
        print(f"  SUM range    : {sum_range}")
        print(f"  SUM cell     : {sum_cell_addr}  ←  {sum_formula}")

        # ── Step 5: Place SUM formula in score column at anchor row ──────────
        # If sum_cell is inside a merged region, unmerge it first
        unmerge_if_needed(ws, sum_cell_addr)
        sum_cell = ws[sum_cell_addr]

        # Preserve font style of the header cell if possible
        try:
            ref_font = ws.cell(header_row, score_col).font
            sum_cell.font = Font(
                name=ref_font.name or 'Angsana New',
                bold=True,
                size=ref_font.size or 14,
                color='FF0000',
            )
        except Exception:
            sum_cell.font = Font(name='Angsana New', bold=True, size=14, color='FF0000')
        sum_cell.alignment = Alignment(horizontal='center', vertical='center')
        sum_cell.value = sum_formula

        # ── Step 6: Update anchor text cell to reference formula ─────────────
        anchor_txt = anchor.value.strip()

        # Skip if anchor already contains a valid formula referencing our sum cell
        if anchor_txt.startswith('=') and sum_cell_addr in anchor_txt:
            print(f"  Anchor cell  : already references {sum_cell_addr} — skipped")
        else:
            # Extract label prefix (everything before any '=' or digit)
            label_match = re.match(r'^([^\d=]+)', anchor_txt)
            label = label_match.group(1).rstrip(' =') if label_match else anchor_txt

            old_merge = unmerge_if_needed(ws, anchor.coordinate)
            anchor_cell = ws[anchor.coordinate]
            safe_label = label.replace('"', '""')
            anchor_cell.value = f'="{safe_label}  = "&TEXT({sum_cell_addr},"0.##")'
            try:
                anchor_cell.font = Font(
                    name=sum_cell.font.name,
                    bold=True,
                    size=sum_cell.font.size,
                    color='FF0000',
                )
                anchor_cell.alignment = Alignment(horizontal='center', vertical='center')
            except Exception:
                pass
            if old_merge:
                ws.merge_cells(old_merge)
            print(f"  Anchor cell  : updated to reference {sum_cell_addr} dynamically")
        results.append({
            'sheet': sheet_name,
            'score_col': col_letter,
            'sum_cell': sum_cell_addr,
            'formula': sum_formula,
            'anchor': anchor.coordinate,
            'confirmed_total': direct_sum,
        })

    if not results:
        print("\n[WARNING] No sheets were fixed — check keyword coverage.")
        return False

    wb.save(output_path)
    print(f"\n{'='*60}")
    print(f"Saved → {output_path}")
    print(f"{'='*60}")
    for r in results:
        print(f"  Sheet '{r['sheet']}': {r['sum_cell']} = {r['formula']}")
        print(f"    Confirmed total = {r['confirmed_total']}")
    return True


# ── CLI entry point ───────────────────────────────────────────────────────────
if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python fix_p4p_score.py <excel_file> [output_file]")
        sys.exit(1)
    inp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) >= 3 else None
    ok = fix_p4p_file(inp, out)
    sys.exit(0 if ok else 1)
