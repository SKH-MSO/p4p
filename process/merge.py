#!/usr/bin/env python3
"""
Excel merge helper — openpyxl-based, preserves original cell formatting.
Usage: python merge.py <manifest.json> <output.xlsx>

manifest.json schema:
{
  "files": [
    {
      "path":       "<absolute path to source .xlsx>",
      "sheet_name": "<destination sheet name, max 31 chars>",
      "tab_color":  "<AARRGGBB hex string, or empty>"
    }
  ]
}

Exit codes:
  0 — success
  1 — bad arguments
  2 — no sheets were copied (all sources empty or unreadable)
"""

import sys
import json
import copy

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    print("ERROR: openpyxl is not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ─── helpers ────────────────────────────────────────────────────────────────

def is_sheet_empty(ws):
    """Return True if the worksheet has 1 or fewer rows that contain any data."""
    data_rows = 0
    for row in ws.iter_rows():
        if any(cell.value is not None and cell.value != '' for cell in row):
            data_rows += 1
            if data_rows > 1:
                return False
    return True  # 0 or 1 rows with data → treat as empty


def copy_sheet(src_ws, dst_wb, sheet_name, tab_color=None):
    """
    Copy src_ws into dst_wb as a new sheet named sheet_name.

    Preserves:
      - Cell values (including formulas)
      - Cell styles: font, fill, border, alignment, number format, protection
      - Merged cell ranges
      - Column widths and row heights (including hidden state)
      - Freeze panes
      - Print settings (page setup, margins, print area, title rows/cols)
      - Sheet view (zoom, grid lines)
      - Conditional formatting rules
      - Tab colour (from argument, not from source)
    """
    dst_ws = dst_wb.create_sheet(title=sheet_name)

    # Tab colour
    if tab_color:
        dst_ws.sheet_properties.tabColor = tab_color

    # ── Column dimensions ────────────────────────────────────────────
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_col = dst_ws.column_dimensions[col_letter]
        dst_col.width    = col_dim.width
        dst_col.hidden   = col_dim.hidden
        dst_col.bestFit  = col_dim.bestFit
        dst_col.outline_level = col_dim.outline_level

    # ── Row dimensions ───────────────────────────────────────────────
    for row_idx, row_dim in src_ws.row_dimensions.items():
        dst_row = dst_ws.row_dimensions[row_idx]
        dst_row.height        = row_dim.height
        dst_row.hidden        = row_dim.hidden
        dst_row.outline_level = row_dim.outline_level

    # ── Merged cells (must come before cell writes) ──────────────────
    for merge_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merge_range))

    # ── Cells: values + styles ───────────────────────────────────────
    # MergedCell objects (non-top-left cells of a merge range) are read-only;
    # skip them — their range is already registered via merge_cells() above.
    from openpyxl.cell.cell import MergedCell
    for row in src_ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column)
            dst_cell.value = cell.value
            if cell.has_style:
                dst_cell.font          = copy.copy(cell.font)
                dst_cell.border        = copy.copy(cell.border)
                dst_cell.fill          = copy.copy(cell.fill)
                dst_cell.number_format = cell.number_format
                dst_cell.protection    = copy.copy(cell.protection)
                dst_cell.alignment     = copy.copy(cell.alignment)

    # ── Page setup / print settings ─────────────────────────────────
    dst_ws.page_setup        = copy.copy(src_ws.page_setup)
    dst_ws.page_margins      = copy.copy(src_ws.page_margins)
    dst_ws.print_title_rows  = src_ws.print_title_rows
    dst_ws.print_title_cols  = src_ws.print_title_cols
    if src_ws.print_area:
        dst_ws.print_area = src_ws.print_area

    # ── Freeze panes ─────────────────────────────────────────────────
    # Use ySplit/xSplit directly rather than freeze_panes (which returns
    # topLeftCell — i.e. where the user had *scrolled* the bottom pane —
    # not the actual freeze boundary).  Using topLeftCell would freeze
    # far too many rows when the bottom pane was scrolled down.
    pane = src_ws.sheet_view.pane
    if pane is not None and pane.state == 'frozen':
        from openpyxl.utils import get_column_letter
        x_split = int(pane.xSplit or 0)
        y_split = int(pane.ySplit or 0)
        if x_split > 0 or y_split > 0:
            col = get_column_letter(x_split + 1)
            row = y_split + 1
            dst_ws.freeze_panes = f'{col}{row}'

    # ── Sheet view (zoom, grid lines, row/col headers) ───────────────
    sv = src_ws.sheet_view
    dst_ws.sheet_view.zoomScale          = sv.zoomScale
    dst_ws.sheet_view.showGridLines      = sv.showGridLines
    dst_ws.sheet_view.showRowColHeaders  = sv.showRowColHeaders

    # ── Conditional formatting ───────────────────────────────────────
    for cf_range, rules in src_ws.conditional_formatting._cf_rules.items():
        for rule in rules:
            dst_ws.conditional_formatting.add(cf_range, rule)

    return dst_ws


# ─── main ────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) != 3:
        print("Usage: python merge.py <manifest.json> <output.xlsx>", file=sys.stderr)
        sys.exit(1)

    manifest_path = sys.argv[1]
    output_path   = sys.argv[2]

    with open(manifest_path, encoding='utf-8') as f:
        manifest = json.load(f)

    dst_wb = Workbook()
    dst_wb.remove(dst_wb.active)   # remove the default empty sheet

    sheet_count = 0
    total_rows  = 0

    for entry in manifest['files']:
        src_path   = entry['path']
        sheet_name = entry['sheet_name']
        tab_color  = entry.get('tab_color') or None

        try:
            src_wb = load_workbook(src_path, data_only=False)
        except Exception as e:
            print(f"  WARNING: cannot open {src_path}: {e}", file=sys.stderr)
            continue

        # Find first non-empty sheet
        target_ws = None
        for ws in src_wb.worksheets:
            if not is_sheet_empty(ws):
                target_ws = ws
                print(f"  -> using sheet \"{ws.title}\" from \"{src_path}\"")
                break
            else:
                print(f"  -> sheet \"{ws.title}\" is empty, trying next")

        if target_ws is None:
            print(f"  WARNING: all sheets empty in {src_path} — skipped", file=sys.stderr)
            continue

        copy_sheet(target_ws, dst_wb, sheet_name, tab_color)
        row_count = target_ws.max_row or 0
        total_rows  += max(0, row_count - 1)
        sheet_count += 1
        print(f"  OK {row_count} rows -> \"{sheet_name}\"")

    if sheet_count == 0:
        print("  ERROR: no sheets were copied", file=sys.stderr)
        sys.exit(2)

    dst_wb.save(output_path)
    print(f"\nSaved {sheet_count} sheets, {total_rows} data rows -> {output_path}")


if __name__ == '__main__':
    main()
